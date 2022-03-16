import openpyxl
import easygui

filename = easygui.fileopenbox('Выбери файл для импорта')

wb = openpyxl.load_workbook(filename)
ws = wb['Лист1']

observation = {}
codes = ''

for i in range(2, ws.max_row):
    codes = codes + '\'' + ws.cell(i, 1).value + '\', '
    observation[ws.cell(i, 1).value] = ws.cell(i, 2).value

with open(filename+'.txt', 'w') as f:
    if(f.write(codes[:-2])):
        easygui.msgbox('Экспортировано в файл: '+filename+'.txt')

#print(observation['10-002'])


